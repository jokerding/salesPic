#!/usr/bin/env python
# -*- coding: utf-8 -*-
# @File  : tools.py
# @Author: joker
# @Date  : 2018/10/23
# @Desc  :
import operator
import os
import sys
from PyQt5.QtWidgets import *
from openpyxl import load_workbook,Workbook
import json

class Tools(object):
    tablename = None
    @staticmethod
    def resource_path(relative_path):
        """定义一个读取相对路径的函数"""

        # print(os.path.abspath(os.path.dirname(os.path.dirname(__file__))))

        # base_path = os.path.abspath(os.path.dirname(os.path.dirname(__file__)))

        if hasattr(sys, "_MEIPASS"):
            base_path = sys._MEIPASS
        else:
            base_path = os.path.abspath(".")

        # print(base_path)
        # print('==============' + os.path.join(base_path, relative_path))
        return os.path.join(base_path, relative_path)

    @staticmethod
    def showAlert(msg):
        if len(msg) == 0:
            msg = "请检查没有传入数据"
        msg_box = QMessageBox(QMessageBox.Warning,"提示", msg)
        msg_box.show()
        msg_box.exec_()

    @staticmethod
    def open(sender=None,callBack=None,parent = None):
        if sender == 1:
            directory1 = QFileDialog.getExistingDirectory(parent,
                                                          "选取文件夹",
                                                  "./")  # 起始路径
            print(directory1)
            return directory1

        elif sender == 2:
            fileName, filetype = QFileDialog.getOpenFileName(parent,
                                                             "选取文件",
                                                             "./",
                                                             "All Files (*);;Text Files (*.txt)")  # 设置文件扩展名过滤,注意用双分号间隔
            print(fileName, filetype)

            return fileName

        elif sender == 3:
            files, ok = QFileDialog.getOpenFileNames(parent,
                                                     "多文件选择",
                                                     "./",
                                                     "All Files (*);;Text Files (*.txt)")
            print(files, ok)
            return files
        else:
            fileName, ok = QFileDialog.getSaveFileName(parent,
                                                       "文件保存",
                                                       "./",
                                                       "All Files (*);;Text Files (*.txt)")
            if ok:
                return True
            else:
                return False
        # Tools.showAlert("123456")

    @staticmethod
    def getDirAllFiles(dirPaht=None):
        fileList = []
        for fpathe,dirs,files in os.walk(dirPaht):
            for file in files:
                fileList.append(os.path.join(fpathe,file))
        return fileList

    @staticmethod
    def writeExcelFile(filePath=None,values=None,sheetTitle=None):
        wb = Workbook()
        sheet = wb.active
        sheet.title = sheetTitle
        for i in range(len(values)):
            sheet.cell(row=1,column=i+1,value=str(values[i]))
        if wb.save(filePath):
            return filePath

        print("写入数据成功")

    @staticmethod
    def writeFiles(dicInfo, path=None):
        if not path:
            path = os.path.join(os.path.abspath('.'), 'config.json')
            print(path)
        with open(path, 'wb+') as f:
            f.write(json.dumps(dicInfo).encode('utf-8'))
            return True

    @staticmethod
    def redFiles(path=None):
        if not path:
            path = os.path.join(os.path.abspath('.'), 'config.json')
            print(path)
        try:
            with open(path) as f:
                res = f.read()
                return json.loads(res)
        except FileNotFoundError:
            return None



    @staticmethod
    def readExcelFiel(filepath=None, dic=None,sheetIndex=None):
        excel = load_workbook(filepath)
        if not sheetIndex:
            sheetIndex = 0
        sheet = excel.worksheets[sheetIndex]
        list =[]
        i = 0
        for row in sheet.rows:
            values = []
            for cell in row:
                if i == 0:
                    pass
                else:
                    values.append(str(cell.value))
            if values:
                values.insert(0,i-1)
                list.append(values)
            i += 1

        print(list)

        return list

        # Tools.insert_into_table(tablename=Tools.tablename,values=list)

    @staticmethod
    def insert_into_table(tablename=None,values=None):
        from src.util.dbUtils import SqliteDbUtil
        if tablename == None:
            tablename = "wx_package"

        for maps in values:
            print(tuple(maps))
            sql = "insert into %s  values %s" % (tablename,tuple(maps))
            print(sql)
            SqliteDbUtil.DML(sql=sql, values=maps)




class mainWindow(QWidget):
    def __init__(self):
        super(mainWindow,self).__init__()
        self.initUI()

    def initUI(self):
        layout = QVBoxLayout()
        self.setLayout(layout)
        self.opepdir = QPushButton('选取文件夹')
        self.opepdir.clicked.connect(lambda :self.open(1))

        self.opepFile = QPushButton('选取文件')
        self.opepFile.clicked.connect(lambda :self.open(2))

        self.opepFiles = QPushButton('选取多个文件')
        self.opepFiles.clicked.connect(lambda :self.open(3))

        self.savaFile = QPushButton('保存文件')
        self.savaFile.clicked.connect(lambda :self.open(4))

        layout.addWidget(self.opepdir)
        layout.addWidget(self.opepFile)
        layout.addWidget(self.opepFiles)
        layout.addWidget(self.savaFile)


    def open(self,sender):
        if sender == 1:
            directory1 = QFileDialog.getExistingDirectory(self,
                                                          "选取文件夹",
                                                          "./")  # 起始路径
            print(directory1)
            return directory1

        elif sender == 2:
            fileName, filetype = QFileDialog.getOpenFileName(self,
                                                             "选取文件",
                                                             "./",
                                                             "All Files (*);;Text Files (*.txt)")  # 设置文件扩展名过滤,注意用双分号间隔
            print(fileName, filetype)

            if fileName:
                self.loadFile(fileName)

            return fileName

        elif sender == 3:
            files, ok = QFileDialog.getOpenFileNames(self,
                                                     "多文件选择",
                                                     "./",
                                                     "All Files (*);;Text Files (*.txt)")
            print(files, ok)
            return files
        else:
            fileName, ok = QFileDialog.getSaveFileName(self,
                                                       "文件保存",
                                                       "./",
                                                       "All Files (*);;Text Files (*.txt)")

            print(fileName)
            if ok:
                return True
            else:
                return False
        # Tools.showAlert("123456")

    def loadFile(self,path):
        Tools.readExcelFiel(path)


# Tools.redFiles()/

# if __name__ == '__main__':
#
#     import sys
#
#     app  = QApplication(sys.argv)
#
#     window = mainWindow()
#
#     window.show()
#
#
#     sys.exit(app.exec_())



# Tools.readExcelFiel("/Users/dingbin/Downloads/wx.xlsx")

# print(Tools.wxPacklist())
